#!/usr/bin/env python3
"""
Анализ Excel файла для извлечения фильтров и изображений
"""
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import os

wb = load_workbook('Расчет_образцов_выставка_Олимпийский_.xlsx')

print("="*80)
print("АНАЛИЗ ДАННЫХ ДЛЯ ФИЛЬТРОВ")
print("="*80)

# Листы для анализа
sheets_to_analyze = ['Двери ', 'диваны', 'столы', 'Системы хранения', 'стеллажи']

# Словари для сбора уникальных значений
colors = set()
materials = set()
styles = set()
sizes = set()
countries = set()

for sheet_name in sheets_to_analyze:
    if sheet_name not in wb.sheetnames:
        continue

    print(f"\n{'='*80}")
    print(f"ЛИСТ: {sheet_name}")
    print(f"{'='*80}")

    ws = wb[sheet_name]

    # Проверяем наличие изображений
    print(f"\nИзображения на листе: {len(ws._images)}")

    # Анализируем данные построчно
    for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_num <= 2 or not row[0]:  # Пропускаем заголовки
            continue

        # Для разных листов данные в разных колонках
        if sheet_name in ['Двери ', 'Перегородки']:
            # Колонка 8: Отделка профиля / кромки / молдинга цвет
            if len(row) > 7 and row[7]:
                color = str(row[7]).strip()
                if color and len(color) < 100:
                    colors.add(color)

            # Колонка 11: Стекло / отделка
            if len(row) > 11 and row[11]:
                material = str(row[11]).strip()
                if material and len(material) < 100:
                    materials.add(material)

        elif sheet_name in ['диваны']:
            # Колонка 4: Отделка (ткани)
            if len(row) > 4 and row[4]:
                material = str(row[4]).strip()
                if material and len(material) < 100:
                    materials.add(material)

        elif sheet_name in ['столы']:
            # Колонка 7: Отделка профиля
            if len(row) > 7 and row[7]:
                color = str(row[7]).strip()
                if color and len(color) < 100:
                    colors.add(color)

            # Колонка 8: Стекло / отделка
            if len(row) > 8 and row[8]:
                material = str(row[8]).strip()
                if material and len(material) < 100:
                    materials.add(material)

print(f"\n{'='*80}")
print("УНИКАЛЬНЫЕ ЗНАЧЕНИЯ ДЛЯ ФИЛЬТРОВ")
print(f"{'='*80}")

print(f"\n--- ЦВЕТА ({len(colors)}) ---")
for color in sorted(colors):
    print(f"  • {color}")

print(f"\n--- МАТЕРИАЛЫ ({len(materials)}) ---")
for material in sorted(materials):
    print(f"  • {material}")

# Дополнительный анализ - извлечение всех изображений
print(f"\n{'='*80}")
print("ИЗВЛЕЧЕНИЕ ИЗОБРАЖЕНИЙ")
print(f"{'='*80}")

image_dir = 'excel_images'
os.makedirs(image_dir, exist_ok=True)

total_images = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if ws._images:
        print(f"\n{sheet_name}: {len(ws._images)} изображений")
        for idx, img in enumerate(ws._images):
            total_images += 1
            # Сохраняем изображение
            img_filename = f"{image_dir}/{sheet_name.strip()}_{idx+1}.{img.format}"
            try:
                with open(img_filename, 'wb') as f:
                    f.write(img._data())
                print(f"  ✓ Сохранено: {img_filename}")
            except Exception as e:
                print(f"  ✗ Ошибка при сохранении {img_filename}: {e}")

print(f"\n\nВСЕГО ИЗОБРАЖЕНИЙ: {total_images}")
print(f"Изображения сохранены в папку: {image_dir}/")
