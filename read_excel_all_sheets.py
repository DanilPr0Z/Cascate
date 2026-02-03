#!/usr/bin/env python3
from openpyxl import load_workbook

wb = load_workbook('Расчет_образцов_выставка_Олимпийский_.xlsx')
print('Листы:', wb.sheetnames)
print('\n' + '='*80 + '\n')

# Читаем несколько интересных листов
sheets_to_read = ['Двери ', 'диваны', 'столы', 'Системы хранения', 'стеллажи']

for sheet_name in sheets_to_read:
    if sheet_name in wb.sheetnames:
        print(f'\n=== ЛИСТ: {sheet_name} ===\n')
        ws = wb[sheet_name]
        print(f'Всего строк: {ws.max_row}')
        print(f'Всего колонок: {ws.max_column}\n')

        # Читаем первые 15 строк
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i <= 15:
                print(f'Строка {i}:', row)

        print('\n' + '-'*80 + '\n')
