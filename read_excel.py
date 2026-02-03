#!/usr/bin/env python3
from openpyxl import load_workbook

wb = load_workbook('Расчет_образцов_выставка_Олимпийский_.xlsx')
print('Листы:', wb.sheetnames)
ws = wb.active
print('\nПервые 30 строк:\n')

for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    print(f'Строка {i}:', row)
    if i >= 30:
        break

print(f'\n\nВсего строк: {ws.max_row}')
print(f'Всего колонок: {ws.max_column}')
