from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse, path
from django.utils.safestring import mark_safe
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.middleware.csrf import get_token
from .models import Category, SubCategory, Product, ProductImage, Store, ProductStock, ProductRating
import os

# Убираем django_cron из админки
try:
    from django_cron.models import CronJobLog, CronJobLock
    admin.site.unregister(CronJobLog)
    admin.site.unregister(CronJobLock)
except Exception:
    pass


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ['image_preview', 'name', 'order', 'get_products_count', 'created_at', 'view_on_site']
    list_editable = ['order']
    prepopulated_fields = {'slug': ('name',)}
    search_fields = ['name']
    list_display_links = ['name']
    list_per_page = 25
    
    fieldsets = (
        ('Основная информация', {
            'fields': ('name', 'slug', 'image', 'description', 'order')
        }),
        ('Метаданные', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    readonly_fields = ['created_at', 'updated_at', 'image_preview']
    
    def image_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" style="width: 60px; height: 60px; object-fit: cover; border-radius: 4px;" />', obj.image.url)
        return "Нет изображения"
    image_preview.short_description = "Изображение"
    
    def view_on_site(self, obj):
        url = obj.get_absolute_url()
        return format_html('<a href="{}" target="_blank">👁️ Просмотр</a>', url)
    view_on_site.short_description = "Просмотр"


@admin.register(SubCategory)
class SubCategoryAdmin(admin.ModelAdmin):
    list_display = ['name', 'category', 'order', 'view_on_site']
    list_editable = ['order']
    prepopulated_fields = {'slug': ('name',)}
    search_fields = ['name', 'category__name']
    list_per_page = 25
    
    def view_on_site(self, obj):
        url = obj.get_absolute_url()
        return format_html('<a href="{}" target="_blank">👁️ Просмотр</a>', url)
    view_on_site.short_description = "Просмотр"



class ProductImageInline(admin.TabularInline):
    model = ProductImage
    extra = 1
    fields = ['image', 'image_preview', 'is_main', 'order', 'alt_text']
    readonly_fields = ['image_preview']

    def image_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" style="width: 100px; height: 100px; object-fit: cover; border-radius: 4px;" />', obj.image.url)
        return "Нет изображения"
    image_preview.short_description = "Превью"


class ProductStockInline(admin.TabularInline):
    model = ProductStock
    extra = 1
    fields = ['store', 'quantity']
    autocomplete_fields = ['store']


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ['image_preview', 'name', 'category', 'price_formatted', 'discount_badge', 'availability_badge', 'is_new_display', 'created_at', 'view_on_site']
    prepopulated_fields = {'slug': ('name',)}
    search_fields = ['name', '=product_number', 'slug']
    actions = ['assign_to_store']
    inlines = [ProductImageInline, ProductStockInline]
    list_per_page = 25
    list_display_links = ['name']
    date_hierarchy = 'created_at'

    def changelist_view(self, request, extra_context=None):
        """Добавляем кнопки импорта в контекст"""
        extra_context = extra_context or {}
        extra_context['download_template_url'] = reverse('admin:catalog_product_download_template')
        extra_context['import_excel_url'] = reverse('admin:catalog_product_import_excel')
        return super().changelist_view(request, extra_context=extra_context)

    fieldsets = (
        ('Основная информация', {
            'fields': ('name', 'slug', 'category', 'subcategory', 'price', 'price_from', 'discount'),
            'classes': ('wide',)
        }),
        ('Характеристики товара', {
            'fields': ('materials', 'dimensions', 'product_number', 'availability', 'tour_3d_url', 'map_point'),
        }),
        ('Описание', {
            'fields': ('short_description', 'description'),
            'classes': ('wide',)
        }),
        ('QR код и карточка товара', {
            'fields': ('qr_code', 'qr_code_preview', 'product_card_preview', 'generate_card_button'),
            'classes': ('collapse',)
        }),
    )
    readonly_fields = ['created_at', 'updated_at', 'views_count', 'image_preview', 'qr_code_preview', 'product_card_preview', 'generate_card_button']

    class Media:
        js = ('admin/js/product_subcategory_filter.js',)

    def get_urls(self):
        """Добавляем URL для генерации карточки и импорта"""
        urls = super().get_urls()
        custom_urls = [
            path(
                '<int:product_id>/generate-card/',
                self.admin_site.admin_view(self.generate_card_view),
                name='catalog_product_generate_card',
            ),
            path(
                'download-template/',
                self.admin_site.admin_view(self.download_template_view),
                name='catalog_product_download_template',
            ),
            path(
                'import-excel/',
                self.admin_site.admin_view(self.import_excel_view),
                name='catalog_product_import_excel',
            ),
            path(
                'assign-to-store/',
                self.admin_site.admin_view(self.assign_to_store_view),
                name='catalog_product_assign_to_store',
            ),
            path(
                'export-excel/',
                self.admin_site.admin_view(self.export_excel_view),
                name='catalog_product_export_excel',
            ),
            path(
                'subcategories-by-category/',
                self.admin_site.admin_view(self.subcategories_by_category_view),
                name='catalog_product_subcategories_by_category',
            ),
        ]
        return custom_urls + urls

    def subcategories_by_category_view(self, request):
        """AJAX: вернуть подкатегории для выбранной категории"""
        category_id = request.GET.get('category_id')
        if not category_id:
            return JsonResponse({'subcategories': []})
        subcategories = SubCategory.objects.filter(category_id=category_id).order_by('order', 'name')
        data = [{'id': s.id, 'name': s.name} for s in subcategories]
        return JsonResponse({'subcategories': data})

    def export_excel_view(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Товары"

        headers = [
            'Название', 'Артикул', 'Материалы', 'Размеры',
            'Цена', 'Цена от', 'Скидка %',
            'Наличие', 'Категория', 'Подкатегория',
            'Описание', 'Ссылка на 3D тур', 'Точка на карте',
        ]
        header_fill = PatternFill(start_color='444444', end_color='444444', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        products = Product.objects.select_related('category', 'subcategory').order_by('name')
        for row, product in enumerate(products, 2):
            ws.cell(row=row, column=1, value=product.name)
            ws.cell(row=row, column=2, value=product.product_number or '')
            ws.cell(row=row, column=3, value=product.materials or '')
            ws.cell(row=row, column=4, value=product.dimensions or '')
            ws.cell(row=row, column=5, value=float(product.price))
            ws.cell(row=row, column=6, value=float(product.price_from) if product.price_from else '')
            ws.cell(row=row, column=7, value=product.discount or 0)
            ws.cell(row=row, column=8, value=product.availability or '')
            ws.cell(row=row, column=9, value=product.category.name if product.category else '')
            ws.cell(row=row, column=10, value=product.subcategory.name if product.subcategory else '')
            ws.cell(row=row, column=11, value=product.description or '')
            ws.cell(row=row, column=12, value=product.tour_3d_url or '')
            ws.cell(row=row, column=13, value=int(product.map_point) if product.map_point and product.map_point.isdigit() else (product.map_point or ''))

        col_widths = [40, 18, 30, 20, 14, 14, 12, 20, 25, 25, 40, 35, 18]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
        wb.save(response)
        return response

    def assign_to_store(self, request, queryset):
        """Действие: перенаправить на страницу выбора магазина"""
        from django.http import HttpResponseRedirect
        selected_ids = ','.join(str(obj.pk) for obj in queryset)
        return HttpResponseRedirect(
            f'/admin/catalog/product/assign-to-store/?ids={selected_ids}'
        )
    assign_to_store.short_description = "Добавить выбранные в магазин"

    def assign_to_store_view(self, request):
        """Промежуточная страница выбора магазина для добавления товаров"""
        from django.http import HttpResponseRedirect
        from django.middleware.csrf import get_token

        if request.method == 'POST':
            ids_raw = request.POST.get('ids', '')
            store_id = request.POST.get('store_id')
            if ids_raw and store_id:
                try:
                    store = Store.objects.get(pk=store_id, is_active=True)
                    product_ids = [int(i) for i in ids_raw.split(',') if i.strip().isdigit()]
                    products = Product.objects.filter(pk__in=product_ids)
                    added = 0
                    for product in products:
                        _, created = ProductStock.objects.get_or_create(
                            product=product,
                            store=store,
                            defaults={'quantity': 1}
                        )
                        if created:
                            added += 1
                    skipped = len(product_ids) - added
                    msg = f'Добавлено {added} товаров в магазин "{store.name}".'
                    if skipped:
                        msg += f' Пропущено {skipped} (уже были в магазине).'
                    self.message_user(request, msg)
                except Store.DoesNotExist:
                    self.message_user(request, 'Магазин не найден.', level='error')
            return HttpResponseRedirect('/admin/catalog/product/')

        # GET — показываем форму
        ids_raw = request.GET.get('ids', '')
        product_ids = [int(i) for i in ids_raw.split(',') if i.strip().isdigit()]
        product_count = len(product_ids)
        stores = Store.objects.filter(is_active=True)
        csrf_token = get_token(request)

        store_options = ''.join(
            f'<option value="{s.pk}">{s.name}</option>' for s in stores
        )

        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Добавить товары в магазин</title>
            <style>
                body {{ font-family: Arial, sans-serif; max-width: 500px; margin: 50px auto; padding: 20px; }}
                h1 {{ color: #444; font-size: 22px; margin-bottom: 20px; }}
                .info {{ background: #f5f5f5; padding: 15px; border-radius: 6px; margin-bottom: 20px; }}
                select {{ width: 100%; padding: 10px; border: 2px solid #ddd; border-radius: 6px;
                          font-size: 14px; margin: 10px 0 20px 0; }}
                button {{ background: #444; color: white; padding: 12px 24px; border: none;
                          border-radius: 4px; cursor: pointer; font-size: 15px; font-weight: 500; }}
                button:hover {{ background: #333; }}
                .back-link {{ display: inline-block; margin-top: 15px; color: #444;
                               text-decoration: none; font-size: 14px; }}
                .back-link:hover {{ text-decoration: underline; }}
            </style>
        </head>
        <body>
            <h1>Добавить товары в магазин</h1>
            <div class="info">Выбрано товаров: <strong>{product_count}</strong></div>
            <form method="post">
                <input type="hidden" name="csrfmiddlewaretoken" value="{csrf_token}">
                <input type="hidden" name="ids" value="{ids_raw}">
                <label for="store_id"><strong>Выберите магазин:</strong></label>
                <select name="store_id" id="store_id" required>
                    <option value="">— выберите магазин —</option>
                    {store_options}
                </select>
                <button type="submit">Добавить в магазин</button>
            </form>
            <a href="/admin/catalog/product/" class="back-link">← Вернуться к товарам</a>
        </body>
        </html>
        """
        return HttpResponse(html)

    def download_template_view(self, request):
        """Скачать шаблон Excel для импорта товаров"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return HttpResponse(
                'Библиотека openpyxl не установлена. Установите: pip install openpyxl',
                status=500
            )

        # Создаем книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Товары"

        # Заголовки столбцов
        headers = [
            'Название', 'Артикул', 'Материалы', 'Размеры',
            'Цена', 'Цена от', 'Скидка %',
            'Наличие', 'Категория', 'Подкатегория',
            'Описание', 'Ссылка на 3D тур', 'Точка на карте',
        ]

        # Стиль заголовков
        header_fill = PatternFill(start_color="20B2AA", end_color="20B2AA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Записываем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Добавляем пример строки
        ws.cell(row=2, column=1, value='ОБЕДЕННЫЙ СТОЛ')
        ws.cell(row=2, column=2, value='ART-001')
        ws.cell(row=2, column=3, value='Дерево, металл')
        ws.cell(row=2, column=4, value='120x80x75 см')
        ws.cell(row=2, column=5, value=15000)
        ws.cell(row=2, column=6, value=12000)
        ws.cell(row=2, column=7, value=10)
        ws.cell(row=2, column=8, value='in_stock')
        ws.cell(row=2, column=9, value='Столовые')
        ws.cell(row=2, column=10, value='Столы обеденные')
        ws.cell(row=2, column=11, value='Описание товара')
        ws.cell(row=2, column=12, value='https://example.com/3d-tour')
        ws.cell(row=2, column=13, value=1)

        # Инструкция на втором листе
        ws2 = wb.create_sheet("Инструкция")
        ws2.cell(row=1, column=1, value="ИНСТРУКЦИЯ ПО ЗАПОЛНЕНИЮ")
        ws2.cell(row=1, column=1).font = Font(bold=True, size=14)

        instructions = [
            "",
            "1. Название - обязательное поле, название товара",
            "2. Артикул - номер товара (опционально)",
            "3. Материалы - из чего сделан товар (опционально)",
            "4. Размеры - габариты товара (опционально)",
            "5. Цена - цена товара в рублях (обязательно)",
            "6. Цена от - минимальная цена 'от' (опционально)",
            "7. Скидка % - скидка в процентах 0-99 (опционально, 0 = нет скидки)",
            "8. Наличие - статус наличия (опционально):",
            "   in_stock = В наличии",
            "   on_the_way = Товар в пути",
            "   new_2025 = Новинка 2025 года",
            "   new_arrival = Новое поступление",
            "   best_offer = Лучшее предложение",
            "   online_fitting = Онлайн-примерка",
            "9. Категория - название категории (должна существовать в системе)",
            "10. Подкатегория - название подкатегории (опционально)",
            "11. Описание - подробное описание товара (опционально)",
            "12. Ссылка на 3D тур - URL ссылка на 3D тур товара (опционально)",
            "13. Точка на карте - число от 1 до бесконечности (не отображается на сайте)",
            "",
            "Доступные категории и подкатегории:",
        ]

        # Добавляем список категорий с подкатегориями
        categories = Category.objects.prefetch_related('subcategories').all()
        for cat in categories:
            instructions.append(f"  {cat.name}:")
            subcats = cat.subcategories.all()
            if subcats:
                for subcat in subcats:
                    instructions.append(f"    - {subcat.name}")
            else:
                instructions.append(f"    (нет подкатегорий)")
            instructions.append("")

        for row_num, instruction in enumerate(instructions, 3):
            ws2.cell(row=row_num, column=1, value=instruction)

        # Ширина столбцов
        col_widths = [40, 18, 30, 20, 14, 14, 12, 20, 25, 25, 40, 35, 18]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        ws2.column_dimensions['A'].width = 60

        # Создаем HTTP ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=shablon_tovarov.xlsx'

        # Сохраняем в response
        wb.save(response)

        return response

    def import_excel_view(self, request):
        """Импорт товаров из Excel"""
        if request.method == 'POST':
            try:
                from openpyxl import load_workbook
            except ImportError:
                return HttpResponse(
                    'Библиотека openpyxl не установлена. Установите: pip install openpyxl',
                    status=500
                )

            excel_file = request.FILES.get('excel_file')

            if not excel_file:
                return HttpResponse('Файл не загружен', status=400)

            try:
                wb = load_workbook(excel_file)
                ws = wb.active

                created_count = 0
                errors = []

                # Пропускаем заголовок, читаем со 2-й строки
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    if not row[0]:  # Пропускаем пустые строки
                        continue

                    try:
                        # Колонки: Название, Артикул, Материалы, Размеры, Цена, Цена от, Скидка%,
                        #           Наличие, Категория, Подкатегория, Описание, Ссылка 3D, Точка на карте
                        name = row[0]
                        product_number = row[1] if len(row) > 1 and row[1] else ''
                        materials = row[2] if len(row) > 2 and row[2] else ''
                        dimensions = row[3] if len(row) > 3 and row[3] else ''
                        price = row[4] if len(row) > 4 and row[4] else None
                        price_from = row[5] if len(row) > 5 and row[5] else None
                        discount = int(row[6]) if len(row) > 6 and row[6] else 0
                        availability = row[7] if len(row) > 7 and row[7] else None
                        category_name = row[8] if len(row) > 8 and row[8] else None
                        subcategory_name = row[9] if len(row) > 9 and row[9] else None
                        description = row[10] if len(row) > 10 and row[10] else ''
                        tour_3d_url = row[11] if len(row) > 11 and row[11] else ''
                        map_point = str(int(row[12])) if len(row) > 12 and row[12] else ''

                        if not price:
                            errors.append(f"Строка {row_num}: Не указана цена")
                            continue

                        # Находим категорию
                        category = None
                        if category_name:
                            try:
                                category = Category.objects.get(name=category_name)
                            except Category.DoesNotExist:
                                errors.append(f"Строка {row_num}: Категория '{category_name}' не найдена")
                                continue
                        else:
                            errors.append(f"Строка {row_num}: Не указана категория")
                            continue

                        # Находим подкатегорию если указана
                        subcategory = None
                        if subcategory_name:
                            try:
                                subcategory = SubCategory.objects.get(name=subcategory_name, category=category)
                            except SubCategory.DoesNotExist:
                                errors.append(f"Строка {row_num}: Подкатегория '{subcategory_name}' не найдена (товар создан без подкатегории)")

                        # Создаем товар
                        product = Product.objects.create(
                            name=name,
                            product_number=product_number,
                            category=category,
                            subcategory=subcategory,
                            price=price,
                            price_from=price_from if price_from else None,
                            discount=discount,
                            availability=availability or '',
                            materials=materials,
                            dimensions=dimensions,
                            description=description,
                            tour_3d_url=tour_3d_url,
                            map_point=map_point,
                        )

                        created_count += 1

                    except Exception as e:
                        errors.append(f"Строка {row_num}: {str(e)}")

                # Формируем ответ
                result = f"Успешно создано товаров: {created_count}"
                if errors:
                    result += "\n\nОшибки:\n" + "\n".join(errors)

                return HttpResponse(f"<pre>{result}</pre><br><a href='/admin/catalog/product/'>Вернуться к товарам</a>")

            except Exception as e:
                return HttpResponse(f'Ошибка при обработке файла: {str(e)}', status=500)

        # GET запрос - показываем форму
        csrf_token = get_token(request)
        template_url = reverse('admin:catalog_product_download_template')

        html = f"""
            <html>
            <head>
                <title>Импорт товаров</title>
                <style>
                    body {{ font-family: Arial, sans-serif; max-width: 600px; margin: 50px auto; padding: 20px; }}
                    h1 {{ color: #20B2AA; }}
                    .upload-form {{ background: #f5f5f5; padding: 30px; border-radius: 8px; }}
                    input[type="file"] {{ margin: 20px 0; }}
                    button {{ background: #20B2AA; color: white; padding: 12px 24px; border: none; border-radius: 4px; cursor: pointer; font-size: 16px; }}
                    button:hover {{ background: #1a9a92; }}
                    .back-link {{ display: inline-block; margin-top: 20px; color: #20B2AA; text-decoration: none; }}
                </style>
            </head>
            <body>
                <h1>📥 Импорт товаров из Excel</h1>
                <div class="upload-form">
                    <p><strong>Инструкция:</strong></p>
                    <ol>
                        <li>Скачайте шаблон Excel</li>
                        <li>Заполните данные о товарах</li>
                        <li>Загрузите заполненный файл</li>
                    </ol>
                    <form method="post" enctype="multipart/form-data">
                        <input type="hidden" name="csrfmiddlewaretoken" value="{csrf_token}">
                        <input type="file" name="excel_file" accept=".xlsx" required>
                        <br>
                        <button type="submit">Загрузить и импортировать</button>
                    </form>
                    <a href="{template_url}" class="back-link">📄 Скачать шаблон Excel</a>
                    <br>
                    <a href="/admin/catalog/product/" class="back-link">← Вернуться к товарам</a>
                </div>
            </body>
            </html>
        """
        return HttpResponse(html)

    def generate_card_view(self, request, product_id):
        """View для генерации карточки товара через AJAX"""
        import traceback
        try:
            product = Product.objects.get(pk=product_id)

            # Импортируем функцию генерации из команды
            try:
                from PIL import Image, ImageDraw, ImageFont
            except ImportError as e:
                return JsonResponse({'success': False, 'error': f'Pillow не установлен: {str(e)}. Установите: pip install Pillow'}, status=500)

            try:
                import qrcode
            except ImportError as e:
                return JsonResponse({'success': False, 'error': f'qrcode не установлен: {str(e)}. Установите: pip install qrcode'}, status=500)

            # Создаем директорию для карточек
            cards_dir = os.path.join(settings.MEDIA_ROOT, 'product_cards')
            os.makedirs(cards_dir, exist_ok=True)

            # Генерируем карточку
            card_path = self._generate_product_card(product, cards_dir)

            return JsonResponse({
                'success': True,
                'message': 'Карточка товара успешно создана',
                'card_url': os.path.join(settings.MEDIA_URL, 'product_cards', f'card_{product.slug}.png')
            })
        except Product.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Товар не найден'}, status=404)
        except Exception as e:
            error_trace = traceback.format_exc()
            print(f"[ERROR] Failed to generate card: {error_trace}")
            return JsonResponse({'success': False, 'error': f'{str(e)}'}, status=500)

    def _generate_product_card(self, product, output_dir):
        """Генерирует карточку товара через label_generator."""
        from catalog.label_generator import generate_product_card
        return generate_product_card(product, output_dir)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """Фильтрация подкатегорий по выбранной категории"""
        if db_field.name == "subcategory":
            # Получаем ID объекта из URL для редактирования
            object_id = request.resolver_match.kwargs.get('object_id')
            if object_id:
                try:
                    product = Product.objects.get(pk=object_id)
                    kwargs["queryset"] = SubCategory.objects.filter(category=product.category)
                except Product.DoesNotExist:
                    pass
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    
    def image_preview(self, obj):
        main_image = obj.get_main_image()
        if main_image:
            return format_html('<img src="{}" style="width: 60px; height: 60px; object-fit: cover; border-radius: 4px;" />', main_image.image.url)
        return "Нет изображения"
    image_preview.short_description = "Изображение"
    
    def price_formatted(self, obj):
        if obj.discount:
            discounted = obj.price * (100 - obj.discount) / 100
            return format_html(
                '<span style="text-decoration:line-through;color:#999;font-size:11px;">{} ₽</span>'
                '&nbsp;<strong style="color:#e53935;">{} ₽</strong>',
                '{:,}'.format(int(obj.price)).replace(',', '\u202f'),
                '{:,}'.format(int(discounted)).replace(',', '\u202f'),
            )
        return format_html(
            '<strong style="color: #20B2AA;">{} ₽</strong>',
            '{:,}'.format(int(obj.price)).replace(',', '\u202f'),
        )
    price_formatted.short_description = "Цена"
    price_formatted.admin_order_field = 'price'

    def discount_badge(self, obj):
        if obj.discount:
            return format_html(
                '<span style="background:#e53935;color:white;padding:3px 8px;border-radius:10px;font-size:11px;font-weight:600;">-{}%</span>',
                obj.discount,
            )
        return '—'
    discount_badge.short_description = "Скидка"
    discount_badge.admin_order_field = 'discount'
    
    def availability_badge(self, obj):
        colors = {
            'in_stock': '#28a745',
            'on_the_way': '#ffc107',
            'new_2025': '#dc3545',
            'new_arrival': '#17a2b8',
            'best_offer': '#6610f2',
            'online_fitting': '#e83e8c',
        }
        color = colors.get(obj.availability, '#6c757d')
        label = dict(Product.AVAILABILITY_CHOICES).get(obj.availability, obj.availability)
        return format_html(
            '<span style="background: {}; color: white; padding: 4px 8px; border-radius: 12px; font-size: 11px; font-weight: 500;">{}</span>',
            color, label
        )
    availability_badge.short_description = "Наличие"
    availability_badge.admin_order_field = 'availability'
    
    def is_new_display(self, obj):
        if obj.is_new:
            return format_html('<span style="background: #DC143C; color: white; padding: 4px 8px; border-radius: 12px; font-size: 11px; font-weight: 500;">NEW</span>')
        return format_html('<span style="color: #999;">—</span>')
    is_new_display.short_description = "Новинка"
    
    def qr_code_preview(self, obj):
        if obj.qr_code:
            return format_html('<img src="{}" style="width: 150px; height: 150px;" />', obj.qr_code.url)
        return "QR код будет сгенерирован автоматически при сохранении"
    qr_code_preview.short_description = "Превью QR кода"

    def product_card_preview(self, obj):
        """Показывает превью карточки товара"""
        if obj.pk:
            # Путь к карточке товара
            card_path = os.path.join(settings.MEDIA_ROOT, 'product_cards', f'card_{obj.slug}.png')

            if os.path.exists(card_path):
                card_url = os.path.join(settings.MEDIA_URL, 'product_cards', f'card_{obj.slug}.png')
                return format_html(
                    '<div style="border: 1px solid #ddd; padding: 10px; display: inline-block; border-radius: 4px;">'
                    '<img src="{}" style="max-width: 400px; height: auto; display: block;" />'
                    '<a href="{}" download="card_{}.png" style="display: inline-block; margin-top: 10px; padding: 8px 16px; background: #20B2AA; color: white; text-decoration: none; border-radius: 4px;">📥 Скачать карточку</a>'
                    '</div>',
                    card_url, card_url, obj.slug
                )
            else:
                return format_html(
                    '<div style="color: #999; padding: 10px; border: 1px dashed #ddd; border-radius: 4px;">'
                    'Карточка товара не создана<br>'
                    '<small>Нажмите кнопку "Сгенерировать карточку" ниже</small>'
                    '</div>'
                )
        return "Сохраните товар для создания карточки"
    product_card_preview.short_description = "Карточка товара"

    def generate_card_button(self, obj):
        """Кнопка для генерации карточки товара"""
        if obj.pk:
            return format_html(
                '<button type="button" onclick="generateProductCard({})" '
                'style="padding: 10px 20px; background: #20B2AA; color: white; border: none; '
                'border-radius: 4px; cursor: pointer; font-size: 14px; font-weight: 500;">'
                '🎴 Сгенерировать карточку товара'
                '</button>'
                '<div id="card-status-{}" style="margin-top: 10px; padding: 10px; border-radius: 4px; display: none;"></div>'
                '<script>'
                'function generateProductCard(productId) {{'
                '  const statusDiv = document.getElementById("card-status-" + productId);'
                '  statusDiv.style.display = "block";'
                '  statusDiv.style.background = "#fff3cd";'
                '  statusDiv.style.color = "#856404";'
                '  statusDiv.innerHTML = "⏳ Генерация карточки...";'
                '  '
                '  fetch("/admin/catalog/product/" + productId + "/generate-card/", {{'
                '    method: "POST",'
                '    headers: {{'
                '      "X-CSRFToken": document.querySelector("[name=csrfmiddlewaretoken]").value'
                '    }}'
                '  }})'
                '  .then(response => response.json())'
                '  .then(data => {{'
                '    if (data.success) {{'
                '      statusDiv.style.background = "#d4edda";'
                '      statusDiv.style.color = "#155724";'
                '      statusDiv.innerHTML = "✅ Карточка успешно создана! Обновите страницу для просмотра.";'
                '      setTimeout(() => window.location.reload(), 1500);'
                '    }} else {{'
                '      statusDiv.style.background = "#f8d7da";'
                '      statusDiv.style.color = "#721c24";'
                '      statusDiv.innerHTML = "❌ Ошибка: " + data.error;'
                '    }}'
                '  }})'
                '  .catch(error => {{'
                '    statusDiv.style.background = "#f8d7da";'
                '    statusDiv.style.color = "#721c24";'
                '    statusDiv.innerHTML = "❌ Ошибка сети: " + error;'
                '  }});'
                '}}'
                '</script>',
                obj.pk, obj.pk
            )
        return "Сохраните товар для генерации карточки"
    generate_card_button.short_description = "Генерация карточки"

    def view_on_site(self, obj):
        url = obj.get_absolute_url()
        return format_html('<a href="{}" target="_blank" style="color: #20B2AA; font-weight: 500;">👁️ Просмотр</a>', url)
    view_on_site.short_description = "Просмотр"

    class Media:
        css = {
            'all': ('admin/css/admin.css',)
        }


@admin.register(Store)
class StoreAdmin(admin.ModelAdmin):
    list_display = ['name', 'address', 'phone', 'is_active', 'products_count']
    search_fields = ['name', 'address', 'phone']
    list_editable = ['is_active']
    list_per_page = 25

    fieldsets = (
        ('Основная информация', {
            'fields': ('name', 'address', 'phone', 'email', 'add_all_products_button'),
        }),
        ('Дополнительно', {
            'fields': ('working_hours', 'is_active'),
        }),
        ('Метаданные', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    readonly_fields = ['created_at', 'updated_at', 'add_all_products_button']

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                '<int:store_id>/add-all-products/',
                self.admin_site.admin_view(self.add_all_products_view),
                name='catalog_store_add_all_products',
            ),
        ]
        return custom_urls + urls

    def add_all_products_view(self, request, store_id):
        """Добавить все товары в магазин"""
        from django.http import HttpResponseRedirect
        from django.middleware.csrf import get_token

        try:
            store = Store.objects.get(pk=store_id)
        except Store.DoesNotExist:
            self.message_user(request, 'Магазин не найден.', level='error')
            return HttpResponseRedirect('/admin/catalog/store/')

        if request.method == 'POST':
            all_products = Product.objects.all()
            added = 0
            for product in all_products:
                _, created = ProductStock.objects.get_or_create(
                    product=product,
                    store=store,
                    defaults={'quantity': 1}
                )
                if created:
                    added += 1
            skipped = all_products.count() - added
            msg = f'Добавлено {added} товаров в магазин "{store.name}".'
            if skipped:
                msg += f' Пропущено {skipped} (уже были в магазине).'
            self.message_user(request, msg)
            return HttpResponseRedirect(
                reverse('admin:catalog_store_change', args=[store_id])
            )

        # GET — страница подтверждения
        total = Product.objects.count()
        csrf_token = get_token(request)
        action_url = reverse('admin:catalog_store_add_all_products', args=[store_id])
        back_url = reverse('admin:catalog_store_change', args=[store_id])

        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Добавить все товары в магазин</title>
            <style>
                body {{ font-family: Arial, sans-serif; max-width: 500px; margin: 50px auto; padding: 20px; }}
                h1 {{ color: #444; font-size: 22px; margin-bottom: 20px; }}
                .info {{ background: #f5f5f5; padding: 15px; border-radius: 6px; margin-bottom: 20px; }}
                button {{ background: #444; color: white; padding: 12px 24px; border: none;
                          border-radius: 4px; cursor: pointer; font-size: 15px; font-weight: 500; }}
                button:hover {{ background: #333; }}
                .back-link {{ display: inline-block; margin-top: 15px; color: #444;
                               text-decoration: none; font-size: 14px; margin-left: 15px; }}
                .back-link:hover {{ text-decoration: underline; }}
            </style>
        </head>
        <body>
            <h1>Добавить все товары в магазин</h1>
            <div class="info">
                Магазин: <strong>{store.name}</strong><br>
                Будет добавлено: <strong>{total} товаров</strong> (уже существующие будут пропущены)
            </div>
            <form method="post" action="{action_url}">
                <input type="hidden" name="csrfmiddlewaretoken" value="{csrf_token}">
                <button type="submit">Добавить все {total} товаров</button>
            </form>
            <a href="{back_url}" class="back-link">← Отмена</a>
        </body>
        </html>
        """
        return HttpResponse(html)

    def add_all_products_button(self, obj):
        if obj.pk:
            url = reverse('admin:catalog_store_add_all_products', args=[obj.pk])
            return format_html(
                '<a href="{}" style="display: inline-block; padding: 8px 16px; background: #444; '
                'color: white; text-decoration: none; border-radius: 4px; font-size: 13px; '
                'font-weight: 500;">Добавить все товары в этот магазин</a>',
                url
            )
        return "Сохраните магазин для использования этой функции"
    add_all_products_button.short_description = "Добавить все товары"

    def products_count(self, obj):
        count = obj.stock.count()
        return format_html('<span style="background: #20B2AA; color: white; padding: 4px 8px; border-radius: 12px; font-size: 12px;">{} товаров</span>', count)
    products_count.short_description = "Товаров в наличии"


@admin.register(ProductStock)
class ProductStockAdmin(admin.ModelAdmin):
    list_display = ['product', 'store', 'quantity', 'updated_at']
    search_fields = ['product__name', 'store__name']
    autocomplete_fields = ['product', 'store']
    list_per_page = 50


@admin.register(ProductRating)
class ProductRatingAdmin(admin.ModelAdmin):
    list_display = ['product', 'rating', 'session_key_short', 'created_at']
    search_fields = ['product__name', 'session_key']
    readonly_fields = ['created_at']
    list_per_page = 50

    def session_key_short(self, obj):
        return f"{obj.session_key[:10]}..." if len(obj.session_key) > 10 else obj.session_key
    session_key_short.short_description = "Ключ сессии"

