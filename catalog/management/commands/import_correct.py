from django.core.management.base import BaseCommand
from catalog.models import Category, SubCategory, Product, ProductImage
import openpyxl
import os
from django.core.files import File
from pathlib import Path


class Command(BaseCommand):
    help = 'Правильный импорт товаров из Excel с корректными названиями'

    def handle(self, *args, **options):
        excel_path = '/Users/danil/PycharmProjects/Mebel/Расчет_образцов_выставка_Олимпийский_.xlsx'
        images_dir = '/Users/danil/PycharmProjects/Mebel/excel_images'

        self.stdout.write('Загрузка Excel файла...')
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # Словарь для маппинга листов на категории
        sheets_mapping = {
            'Двери ': ('Двери и перегородки', 'Двери'),
            'Перегородки': ('Двери и перегородки', 'Перегородки'),
            'Системы хранения': ('Системы хранения', 'Гардеробные'),
            'стеллажи': ('Системы хранения', 'Стеллажи'),
            'витрины': ('Системы хранения', 'Витрины'),
            'столы': ('Столы', 'Столы'),
            'диваны': ('Мягкая мебель', 'Диваны'),
            'кровати ': ('Мягкая мебель', 'Кровати'),
            'полки подвесные ': ('Декор', 'Полки'),
            'стеновые панели ': ('Декор', 'Стеновые панели'),
        }

        total_imported = 0

        for sheet_name, (category_name, subcategory_name) in sheets_mapping.items():
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f'⚠ Лист "{sheet_name}" не найден'))
                continue

            self.stdout.write(f'\n📄 Обработка листа: {sheet_name}')

            # Получаем или создаем категорию
            category, _ = Category.objects.get_or_create(name=category_name)

            # Получаем или создаем подкатегорию если нужна
            subcategory = None
            if subcategory_name:
                subcategory, _ = SubCategory.objects.get_or_create(
                    category=category,
                    name=subcategory_name
                )

            ws = wb[sheet_name]
            imported = self.import_sheet(ws, sheet_name, category, subcategory, images_dir)
            total_imported += imported

        self.stdout.write(self.style.SUCCESS(f'\n✓ Всего импортировано: {total_imported} товаров'))

    def import_sheet(self, ws, sheet_name, category, subcategory, images_dir):
        """Импорт товаров из листа"""
        imported_count = 0
        image_counter = 0  # Счетчик для номера изображения

        # Находим строку с заголовками
        header_row = None
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
            if row and any('Наименование товара' in str(cell) for cell in row if cell):
                header_row = idx
                break

        if not header_row:
            self.stdout.write(self.style.WARNING(f'  Заголовки не найдены'))
            return 0

        # Получаем заголовки
        header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        headers = [str(cell).strip() if cell else '' for cell in header_cells]

        # Определяем индексы нужных колонок
        col_indexes = self.get_column_indexes(headers, sheet_name)

        # Обрабатываем строки с данными
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            if not row or not any(row):
                continue

            # Пропускаем заголовочные строки
            first_cell = str(row[0]).strip() if row[0] else ''
            if first_cell.lower() in ['двери', 'перегородки', 'гардеробные', 'стеллажи', 'стеновые панели', 'полки подвесные', '№ позиции в зале']:
                continue

            try:
                product_data = self.extract_product_data(row, col_indexes, sheet_name)

                if not product_data['name']:
                    continue

                # Создаем товар
                product = self.create_product(product_data, category, subcategory, row_idx)

                # Увеличиваем счетчик изображений
                image_counter += 1

                # Добавляем изображения
                position = str(row[0]).strip() if row[0] else ''
                if position:
                    self.add_images(product, position, images_dir, sheet_name, image_counter)

                imported_count += 1
                self.stdout.write(self.style.SUCCESS(f'  ✓ {product.name}'))

            except Exception as e:
                self.stdout.write(self.style.ERROR(f'  ✗ Ошибка в строке {row_idx}: {str(e)}'))
                continue

        return imported_count

    def get_column_indexes(self, headers, sheet_name):
        """Определяет индексы колонок в зависимости от листа"""
        indexes = {}

        for idx, header in enumerate(headers):
            header_lower = header.lower().strip()

            if 'наименование товара' in header_lower:
                indexes['name'] = idx
            elif 'название' in header_lower and ('профиль' in header_lower or 'конструкция' in header_lower):
                indexes['model'] = idx
            elif 'цена' in header_lower and 'price' not in indexes:
                # Берем ПЕРВУЮ колонку с "цена"
                indexes['price'] = idx
            elif 'размер' in header_lower and 'size' not in indexes:
                # Берем первую колонку с "размер"
                indexes['size'] = idx
            elif 'отделка' in header_lower and ('профиля' in header_lower or 'кромки' in header_lower):
                indexes['material'] = idx
            elif 'стекло' in header_lower or (header_lower == 'отделка' and 'glass' not in indexes):
                indexes['glass'] = idx
            elif 'комментарий' in header_lower:
                indexes['comment'] = idx

        # Отладочная информация
        print(f"  Индексы для листа {sheet_name}: {indexes}")

        return indexes

    def extract_product_data(self, row, col_indexes, sheet_name):
        """Извлекает данные товара из строки"""
        data = {}

        # Базовое название
        base_name_val = row[col_indexes.get('name', 2)] if col_indexes.get('name') is not None else None
        base_name = str(base_name_val).strip() if base_name_val and base_name_val != 'None' else ''

        model_name_val = row[col_indexes.get('model', 3)] if col_indexes.get('model') is not None else None
        model_name = str(model_name_val).strip() if model_name_val and model_name_val != 'None' else ''

        # Формируем правильное название в зависимости от типа листа
        if sheet_name in ['Двери ', 'Перегородки', 'столы']:
            # Для дверей, перегородок и столов: Тип + Модель
            if base_name and model_name:
                data['name'] = f'{base_name} {model_name}'
            else:
                data['name'] = base_name
        else:
            # Для остальных: используем название как есть
            data['name'] = base_name

        # Цена
        price_val = row[col_indexes.get('price', 9)] if col_indexes.get('price') else None
        try:
            if price_val:
                # Если цена в виде строки "132804/ 129626", берем первую
                price_str = str(price_val).strip().split('/')[0].strip()
                data['price'] = float(price_str) if price_str else 0
            else:
                data['price'] = 0
        except:
            data['price'] = 0

        # Размеры
        size_val = row[col_indexes.get('size', 7)] if col_indexes.get('size') else None
        data['dimensions'] = str(size_val).strip() if size_val and size_val != 'None' else ''

        # Материалы
        material_val = row[col_indexes.get('material')] if col_indexes.get('material') and row[col_indexes.get('material')] else None
        glass_val = row[col_indexes.get('glass')] if col_indexes.get('glass') and row[col_indexes.get('glass')] else None

        materials = []
        if material_val and str(material_val).strip() != 'None':
            materials.append(str(material_val).strip())
        if glass_val and str(glass_val).strip() != 'None':
            materials.append(str(glass_val).strip())

        data['materials'] = ' / '.join(materials) if materials else ''

        # Комментарий
        comment_val = row[col_indexes.get('comment')] if col_indexes.get('comment') and row[col_indexes.get('comment')] else None
        data['description'] = str(comment_val).strip() if comment_val and comment_val != 'None' else ''

        # Номер позиции
        position = str(row[0]).strip() if row[0] else ''
        data['product_number'] = position

        return data

    def create_product(self, product_data, category, subcategory, row_idx):
        """Создает товар в базе данных"""
        # Используем product_number как уникальный ключ
        if product_data['product_number']:
            product, created = Product.objects.update_or_create(
                product_number=product_data['product_number'],
                defaults={
                    'name': product_data['name'],
                    'category': category,
                    'subcategory': subcategory,
                    'price': product_data['price'],
                    'dimensions': product_data['dimensions'][:200] if product_data['dimensions'] else '',
                    'materials': product_data['materials'][:500] if product_data['materials'] else '',
                    'description': product_data['description'],
                    'availability': 'in_stock',
                }
            )
        else:
            # Если нет product_number, используем name и category
            product, created = Product.objects.update_or_create(
                name=product_data['name'],
                category=category,
                defaults={
                    'subcategory': subcategory,
                    'price': product_data['price'],
                    'dimensions': product_data['dimensions'][:200] if product_data['dimensions'] else '',
                    'materials': product_data['materials'][:500] if product_data['materials'] else '',
                    'description': product_data['description'],
                    'product_number': product_data['product_number'],
                    'availability': 'in_stock',
                }
            )
        return product

    def add_images(self, product, position, images_dir, sheet_name, image_number):
        """Добавляет изображения к товару"""
        # Удаляем старые изображения
        product.images.all().delete()

        # Ищем файлы изображений
        if not os.path.exists(images_dir):
            return

        # Определяем префикс файла в зависимости от листа
        prefix_mapping = {
            'Двери ': 'Двери',
            'Перегородки': 'Перегородки',
            'стеллажи': 'стеллажи',
            'витрины': 'витрины',
            'столы': 'столы',
            'диваны': 'диваны',
            'кровати ': 'кровати',
            'полки подвесные ': 'полки подвесные',
            'стеновые панели ': 'стеновые панели',
            'Системы хранения': 'Системы хранения',
        }

        prefix = prefix_mapping.get(sheet_name, sheet_name.strip())

        images_found = []
        for file in os.listdir(images_dir):
            if file.startswith('.'):
                continue
            # Ищем файлы вида "Префикс_номер.расширение"
            # Разбираем имя файла
            parts = file.split('_')
            if len(parts) >= 2 and parts[0] == prefix:
                # Извлекаем номер из второй части (до точки)
                number_part = parts[1].split('.')[0]
                try:
                    if int(number_part) == image_number:
                        images_found.append(file)
                except ValueError:
                    continue

        # Сортируем файлы
        images_found.sort()

        # Добавляем изображения
        for idx, image_file in enumerate(images_found):
            image_path = os.path.join(images_dir, image_file)
            try:
                with open(image_path, 'rb') as f:
                    ProductImage.objects.create(
                        product=product,
                        image=File(f, name=image_file),
                        is_main=(idx == 0),
                        order=idx,
                        alt_text=product.name
                    )
            except Exception as e:
                self.stdout.write(self.style.WARNING(f'    Ошибка загрузки {image_file}: {str(e)}'))
