from django.core.management.base import BaseCommand
from catalog.models import Category, SubCategory, Product, ProductImage
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import os
from django.core.files import File
from django.core.files.base import ContentFile
from pathlib import Path
from io import BytesIO
from PIL import Image as PILImage


class Command(BaseCommand):
    help = 'Импорт товаров из Excel с извлечением изображений напрямую из файла'

    def add_arguments(self, parser):
        parser.add_argument(
            '--verbose',
            action='store_true',
            help='Подробный вывод каждого этапа обработки товара',
        )

    def handle(self, *args, **options):
        self.verbose = options.get('verbose', False)
        excel_path = '/Users/danil/PycharmProjects/Mebel/Расчет_образцов_выставка_Олимпийский_.xlsx'

        self.stdout.write('Загрузка Excel файла...')
        wb = openpyxl.load_workbook(excel_path)

        # Словарь для маппинга листов на категории
        sheets_mapping = {
            'Двери ': ('Двери и перегородки', 'Двери'),
            'Перегородки': ('Двери и перегородки', 'Перегородки'),
            'Системы хранения': ('Системы хранения', 'Гардеробные'),
            'стеллажи': ('Системы хранения', 'Стеллажи'),
            'витрины': ('Системы хранения', 'Витрины'),
            'столы': ('Столы', 'Столы'),
            'диваны': ('Мягкая мебель', 'Диваны'),
            'кровати ': ('Мягкая мебель', 'Кровати'),
            'полки подвесные ': ('Декор', 'Полки'),
            'стеновые панели ': ('Декор', 'Стеновые панели'),
        }

        total_imported = 0

        for sheet_name, (category_name, subcategory_name) in sheets_mapping.items():
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f'⚠ Лист "{sheet_name}" не найден'))
                continue

            self.stdout.write(f'\n📄 Обработка листа: {sheet_name}')

            # Получаем или создаем категорию
            category, _ = Category.objects.get_or_create(name=category_name)

            # Получаем или создаем подкатегорию если нужна
            subcategory = None
            if subcategory_name:
                subcategory, _ = SubCategory.objects.get_or_create(
                    category=category,
                    name=subcategory_name
                )

            ws = wb[sheet_name]
            imported = self.import_sheet(ws, sheet_name, category, subcategory)
            total_imported += imported

        self.stdout.write(self.style.SUCCESS(f'\n✓ Всего импортировано: {total_imported} товаров'))

    def import_sheet(self, ws, sheet_name, category, subcategory):
        """Импорт товаров из листа"""
        imported_count = 0
        products_without_images = []

        # Находим строку с заголовками
        header_row = None
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
            if row and any('Наименование товара' in str(cell) for cell in row if cell):
                header_row = idx
                break

        if not header_row:
            self.stdout.write(self.style.WARNING(f'  Заголовки не найдены'))
            return 0

        # Получаем заголовки
        header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        headers = [str(cell).strip() if cell else '' for cell in header_cells]

        # Определяем индексы нужных колонок
        col_indexes = self.get_column_indexes(headers, sheet_name)

        # Создаем словарь изображений по строкам и кеш данных изображений
        images_by_row = {}
        image_data_cache = {}  # Кеш для данных изображений по их ref

        for img in ws._images:
            if hasattr(img.anchor, '_from'):
                # Определяем начальную и конечную строку изображения
                from_row = img.anchor._from.row + 1  # +1 потому что openpyxl использует 0-based индексы
                to_row = from_row  # По умолчанию - только одна строка

                # Если есть информация о конечной позиции изображения
                if hasattr(img.anchor, 'to') and hasattr(img.anchor.to, 'row'):
                    to_row = img.anchor.to.row + 1

                # Добавляем изображение ко ВСЕМ строкам, которые оно покрывает
                # Это важно для spanning изображений (которые занимают несколько строк)
                for row_num in range(from_row, to_row + 1):
                    if row_num not in images_by_row:
                        images_by_row[row_num] = []
                    images_by_row[row_num].append(img)

                # Кешируем данные изображения сразу
                img_ref = img.ref if hasattr(img, 'ref') else id(img)
                if img_ref not in image_data_cache:
                    try:
                        data = img._data()
                        if data:
                            image_data_cache[img_ref] = data
                    except Exception as e:
                        self.stdout.write(self.style.WARNING(f'  ⚠ Ошибка кеширования изображения: {str(e)}'))

        # Обрабатываем строки с данными
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            if not row or not any(row):
                continue

            # Пропускаем заголовочные строки
            first_cell = str(row[0]).strip() if row[0] else ''
            if first_cell.lower() in ['двери', 'перегородки', 'гардеробные', 'стеллажи', 'стеновые панели', 'полки подвесные', '№ позиции в зале', 'диваны', 'витрины']:
                continue

            try:
                # ШАГ 1: Извлекаем данные товара из строки
                product_data = self.extract_product_data(row, col_indexes, sheet_name)

                if not product_data['name']:
                    continue

                # ШАГ 2: Создаем товар в базе данных со всей информацией
                product = self.create_product(product_data, category, subcategory, row_idx)

                # ШАГ 3: Ищем изображения для этого товара
                found_images, distance = self.find_images_for_row(row_idx, images_by_row)

                # ШАГ 4: Добавляем изображение к товару (берем последнее, если их несколько)
                if found_images:
                    # Берем только последнее изображение (оно обычно правильное для товара)
                    last_image = [found_images[-1]]
                    self.add_images_from_excel(product, last_image, image_data_cache)
                    distance_info = f' (строка {row_idx}{distance})' if distance else ''
                    self.stdout.write(self.style.SUCCESS(f'  ✓ {product_data["product_number"]}: {product.name} [1 фото]{distance_info}'))
                else:
                    products_without_images.append(product_data['product_number'])
                    self.stdout.write(self.style.WARNING(f'  ⚠ {product_data["product_number"]}: {product.name} [БЕЗ ФОТО]'))

                imported_count += 1
                # ШАГ 5: Товар полностью обработан, переходим к следующему

            except Exception as e:
                self.stdout.write(self.style.ERROR(f'  ✗ Ошибка в строке {row_idx}: {str(e)}'))
                import traceback
                traceback.print_exc()
                continue

        # Выводим отчет о товарах без изображений
        if products_without_images:
            self.stdout.write(self.style.WARNING(f'  Товары без фото ({len(products_without_images)}): {", ".join(products_without_images)}'))

        return imported_count

    def find_images_for_row(self, row_idx, images_by_row):
        """Ищет изображения ТОЛЬКО в текущей строке товара.
        Возвращает кортеж (список изображений, пустая строка)"""
        # Проверяем ТОЛЬКО текущую строку - не ищем в соседних!
        if row_idx in images_by_row:
            return images_by_row[row_idx], ''

        # Если в текущей строке нет изображения - возвращаем пустой список
        return [], ''

    def get_column_indexes(self, headers, sheet_name):
        """Определяет индексы колонок в зависимости от листа"""
        indexes = {}

        for idx, header in enumerate(headers):
            header_lower = header.lower().strip()

            if 'наименование товара' in header_lower:
                indexes['name'] = idx
            elif 'название' in header_lower and ('профиль' in header_lower or 'конструкция' in header_lower):
                indexes['model'] = idx
            elif 'цена' in header_lower and 'price' not in indexes:
                indexes['price'] = idx
            elif 'размер' in header_lower and 'size' not in indexes:
                indexes['size'] = idx
            elif 'отделка' in header_lower and ('профиля' in header_lower or 'кромки' in header_lower):
                indexes['material'] = idx
            elif 'стекло' in header_lower or (header_lower == 'отделка' and 'glass' not in indexes):
                indexes['glass'] = idx
            elif 'комментарий' in header_lower:
                indexes['comment'] = idx

        return indexes

    def extract_product_data(self, row, col_indexes, sheet_name):
        """Извлекает данные товара из строки"""
        data = {}

        # Базовое название
        base_name_val = row[col_indexes.get('name', 2)] if col_indexes.get('name') is not None else None
        base_name = str(base_name_val).strip() if base_name_val and base_name_val != 'None' else ''

        model_name_val = row[col_indexes.get('model', 3)] if col_indexes.get('model') is not None else None
        model_name = str(model_name_val).strip() if model_name_val and model_name_val != 'None' else ''

        # Формируем правильное название в зависимости от типа листа
        if sheet_name in ['Двери ', 'Перегородки', 'столы']:
            if base_name and model_name:
                data['name'] = f'{base_name} {model_name}'
            else:
                data['name'] = base_name
        else:
            data['name'] = base_name

        # Цена
        price_val = row[col_indexes.get('price', 9)] if col_indexes.get('price') else None
        try:
            if price_val:
                price_str = str(price_val).strip().split('/')[0].strip()
                data['price'] = float(price_str) if price_str else 0
            else:
                data['price'] = 0
        except:
            data['price'] = 0

        # Размеры
        size_val = row[col_indexes.get('size', 7)] if col_indexes.get('size') else None
        data['dimensions'] = str(size_val).strip() if size_val and size_val != 'None' else ''

        # Материалы
        material_val = row[col_indexes.get('material')] if col_indexes.get('material') and col_indexes.get('material') < len(row) and row[col_indexes.get('material')] else None
        glass_val = row[col_indexes.get('glass')] if col_indexes.get('glass') and col_indexes.get('glass') < len(row) and row[col_indexes.get('glass')] else None

        materials = []
        if material_val and str(material_val).strip() != 'None':
            materials.append(str(material_val).strip())
        if glass_val and str(glass_val).strip() != 'None':
            materials.append(str(glass_val).strip())

        data['materials'] = ' / '.join(materials) if materials else ''

        # Комментарий
        comment_val = row[col_indexes.get('comment')] if col_indexes.get('comment') and col_indexes.get('comment') < len(row) and row[col_indexes.get('comment')] else None
        data['description'] = str(comment_val).strip() if comment_val and comment_val != 'None' else ''

        # Номер позиции
        position = str(row[0]).strip() if row[0] else ''
        data['product_number'] = position

        return data

    def create_product(self, product_data, category, subcategory, row_idx):
        """Создает товар в базе данных"""
        if product_data['product_number']:
            product, created = Product.objects.update_or_create(
                product_number=product_data['product_number'],
                defaults={
                    'name': product_data['name'],
                    'category': category,
                    'subcategory': subcategory,
                    'price': product_data['price'],
                    'dimensions': product_data['dimensions'][:200] if product_data['dimensions'] else '',
                    'materials': product_data['materials'][:500] if product_data['materials'] else '',
                    'description': product_data['description'],
                    'availability': 'in_stock',
                }
            )
        else:
            product, created = Product.objects.update_or_create(
                name=product_data['name'],
                category=category,
                defaults={
                    'subcategory': subcategory,
                    'price': product_data['price'],
                    'dimensions': product_data['dimensions'][:200] if product_data['dimensions'] else '',
                    'materials': product_data['materials'][:500] if product_data['materials'] else '',
                    'description': product_data['description'],
                    'product_number': product_data['product_number'],
                    'availability': 'in_stock',
                }
            )
        return product

    def add_images_from_excel(self, product, images_list, image_data_cache):
        """Добавляет изображения к товару из Excel"""
        # Удаляем старые изображения
        product.images.all().delete()

        # Убираем дубликаты изображений по их ref (путь к изображению)
        seen_refs = set()
        unique_images = []
        for img in images_list:
            # Используем ref как уникальный идентификатор изображения
            img_ref = img.ref if hasattr(img, 'ref') else id(img)
            if img_ref not in seen_refs:
                seen_refs.add(img_ref)
                unique_images.append((img, img_ref))

        # Извлекаем данные изображений из кеша
        images_data = []
        for img, img_ref in unique_images:
            if img_ref in image_data_cache:
                images_data.append(image_data_cache[img_ref])
            else:
                self.stdout.write(self.style.WARNING(f'    ⚠ Изображение не найдено в кеше: {img_ref}'))

        # Теперь обрабатываем извлеченные данные
        for idx, image_data in enumerate(images_data):
            try:
                # Открываем изображение через PIL для проверки и конвертации
                pil_image = PILImage.open(BytesIO(image_data))

                # Загружаем изображение полностью в память
                pil_image.load()

                # Конвертируем в RGB если нужно
                if pil_image.mode in ('RGBA', 'LA', 'P'):
                    pil_image = pil_image.convert('RGB')

                # Создаем новый BytesIO для сохранения
                output = BytesIO()
                pil_image.save(output, format='JPEG', quality=85)

                # Получаем bytes из output
                image_content = output.getvalue()

                # Создаем ProductImage с сохранением файла используя ContentFile
                product_image = ProductImage(
                    product=product,
                    is_main=(idx == 0),
                    order=idx,
                    alt_text=product.name
                )
                product_image.image.save(
                    f'{product.product_number}_{idx+1}.jpg',
                    ContentFile(image_content),
                    save=True
                )

            except Exception as e:
                self.stdout.write(self.style.WARNING(f'    Ошибка загрузки изображения {idx+1}: {str(e)}'))
