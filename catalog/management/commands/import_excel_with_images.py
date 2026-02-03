"""
Management команда для импорта данных из Excel файла с изображениями
"""
import os
import shutil
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.utils.text import slugify
from django.core.files import File
from openpyxl import load_workbook
from catalog.models import Category, SubCategory, Product, ProductImage, FilterCategory, FilterValue


class Command(BaseCommand):
    help = 'Импорт данных из Excel файла с изображениями'

    def handle(self, *args, **options):
        excel_file = 'Расчет_образцов_выставка_Олимпийский_.xlsx'

        if not os.path.exists(excel_file):
            self.stdout.write(self.style.ERROR(f'Файл {excel_file} не найден'))
            return

        self.stdout.write(self.style.SUCCESS(f'Загрузка файла {excel_file}...'))
        wb = load_workbook(excel_file)

        # Создаем директорию для медиа если её нет
        media_products = 'media/products'
        os.makedirs(media_products, exist_ok=True)

        # Маппинг листов на категории
        sheets_mapping = {
            'Двери ': 'Двери и перегородки',
            'Перегородки': 'Двери и перегородки',
            'диваны': 'Мягкая мебель',
            'кровати ': 'Мягкая мебель',
            'столы': 'Столы',
            'Системы хранения': 'Системы хранения',
            'стеллажи': 'Системы хранения',
            'витрины': 'Системы хранения',
            'полки подвесные ': 'Декор',
            'стеновые панели ': 'Декор',
        }

        # Подкатегории
        subcategories_mapping = {
            'Двери ': 'Двери',
            'Перегородки': 'Перегородки',
            'диваны': 'Диваны',
            'кровати ': 'Кровати',
            'столы': 'Столы',
            'Системы хранения': 'Гардеробные',
            'стеллажи': 'Стеллажи',
            'витрины': 'Витрины',
            'полки подвесные ': 'Полки',
            'стеновые панели ': 'Стеновые панели',
        }

        created_categories = {}
        created_subcategories = {}
        products_count = 0
        images_count = 0

        # Создаем/получаем категории
        for sheet_name, category_name in sheets_mapping.items():
            if category_name not in created_categories:
                category, created = Category.objects.get_or_create(
                    name=category_name,
                    defaults={'slug': slugify(category_name)}
                )
                created_categories[category_name] = category
                if created:
                    self.stdout.write(self.style.SUCCESS(f'Создана категория: {category_name}'))

        # Получаем фильтры
        try:
            color_category = FilterCategory.objects.get(slug='color')
            material_category = FilterCategory.objects.get(slug='material')
        except FilterCategory.DoesNotExist:
            self.stdout.write(self.style.WARNING('Фильтры не найдены. Запустите create_filters'))
            color_category = None
            material_category = None

        # Импорт данных из каждого листа
        for sheet_name, category_name in sheets_mapping.items():
            if sheet_name not in wb.sheetnames:
                continue

            self.stdout.write(f'\nОбработка листа: {sheet_name}')
            ws = wb[sheet_name]

            category = created_categories[category_name]
            subcategory_name = subcategories_mapping.get(sheet_name)

            # Создаем подкатегорию
            if subcategory_name:
                if subcategory_name not in created_subcategories:
                    subcategory, created = SubCategory.objects.get_or_create(
                        category=category,
                        name=subcategory_name,
                        defaults={'slug': slugify(subcategory_name)}
                    )
                    created_subcategories[subcategory_name] = subcategory
                    if created:
                        self.stdout.write(self.style.SUCCESS(f'  Создана подкатегория: {subcategory_name}'))
                else:
                    subcategory = created_subcategories[subcategory_name]
            else:
                subcategory = None

            # Извлекаем изображения из листа
            sheet_images = {}
            excel_images_dir = 'excel_images'
            for idx, img in enumerate(ws._images):
                img_num = idx + 1
                # Ищем соответствующий файл изображения
                for ext in ['jpeg', 'jpg', 'png']:
                    img_path = f"{excel_images_dir}/{sheet_name.strip()}_{img_num}.{ext}"
                    if os.path.exists(img_path):
                        sheet_images[idx] = img_path
                        break

            # Читаем товары
            image_idx = 0
            for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
                # Пропускаем заголовки и пустые строки
                if row_num <= 2 or not row[0]:
                    continue

                # Извлекаем данные
                position_number = str(row[0]) if row[0] else ''
                product_name = row[2] if len(row) > 2 and row[2] else None
                price = None

                # Ищем цену в последних колонках
                for i in range(len(row) - 1, -1, -1):
                    if row[i] and isinstance(row[i], (int, float)):
                        try:
                            price = Decimal(str(row[i]))
                            if price > 0:
                                break
                        except:
                            pass

                # Пропускаем если нет названия или цены
                if not product_name or not price:
                    continue

                # Создаем товар
                try:
                    full_name = product_name.strip()

                    # Извлекаем размеры и другие данные
                    dimensions = ''
                    materials = ''
                    description = ''

                    # Для разных листов данные в разных колонках
                    if sheet_name in ['Двери ', 'Перегородки']:
                        if len(row) > 12 and row[12]:
                            dimensions = str(row[12])
                        if len(row) > 11 and row[11]:
                            materials = str(row[11])
                        if len(row) > 13 and row[13]:
                            description = str(row[13])
                    elif sheet_name in ['диваны']:
                        if len(row) > 5 and row[5]:
                            dimensions = str(row[5])
                        if len(row) > 4 and row[4]:
                            materials = str(row[4])
                        if len(row) > 8 and row[8]:
                            description = str(row[8])
                    elif sheet_name in ['столы']:
                        if len(row) > 9 and row[9]:
                            dimensions = str(row[9])
                        if len(row) > 8 and row[8]:
                            materials = str(row[8])
                        if len(row) > 12 and row[12]:
                            description = str(row[12])
                    elif sheet_name in ['Системы хранения', 'стеллажи']:
                        if len(row) > 7 and row[7]:
                            dimensions = str(row[7])
                        if len(row) > 6 and row[6]:
                            materials = str(row[6])
                        if len(row) > 10 and row[10]:
                            description = str(row[10])

                    # Создаем уникальный slug
                    base_slug = slugify(full_name)[:490]
                    if position_number:
                        unique_slug = f"{base_slug}-{slugify(position_number)}"[:500]
                    else:
                        unique_slug = base_slug

                    product, created = Product.objects.get_or_create(
                        slug=unique_slug,
                        defaults={
                            'name': full_name,
                            'category': category,
                            'subcategory': subcategory,
                            'price': price,
                            'product_number': position_number,
                            'dimensions': dimensions[:200] if dimensions else '',
                            'materials': materials[:500] if materials else '',
                            'description': description if description else '',
                            'availability': 'in_stock',
                        }
                    )

                    if created:
                        products_count += 1
                        self.stdout.write(f'  ✓ {full_name} - {price} ₽')

                        # Добавляем изображение если есть
                        if image_idx in sheet_images:
                            img_path = sheet_images[image_idx]
                            try:
                                # Копируем изображение в media
                                img_name = f"{product.slug}{os.path.splitext(img_path)[1]}"
                                dest_path = os.path.join(media_products, img_name)
                                shutil.copy(img_path, dest_path)

                                # Создаем запись изображения
                                ProductImage.objects.create(
                                    product=product,
                                    image=f'products/{img_name}',
                                    is_main=True,
                                    order=0
                                )
                                images_count += 1
                                self.stdout.write(f'    📷 Добавлено изображение')
                            except Exception as e:
                                self.stdout.write(self.style.WARNING(f'    ⚠ Ошибка добавления изображения: {e}'))

                        image_idx += 1
                    else:
                        # Обновляем цену если товар уже существует
                        product.price = price
                        product.save()
                        self.stdout.write(f'  ⟳ Обновлен: {full_name}')

                except Exception as e:
                    self.stdout.write(self.style.WARNING(f'  ✗ Ошибка при создании товара "{product_name}": {e}'))

        self.stdout.write(self.style.SUCCESS(f'\n\nИмпорт завершен!'))
        self.stdout.write(self.style.SUCCESS(f'Создано категорий: {len(created_categories)}'))
        self.stdout.write(self.style.SUCCESS(f'Создано подкатегорий: {len(created_subcategories)}'))
        self.stdout.write(self.style.SUCCESS(f'Создано товаров: {products_count}'))
        self.stdout.write(self.style.SUCCESS(f'Добавлено изображений: {images_count}'))
