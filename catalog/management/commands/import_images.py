from django.core.management.base import BaseCommand
from catalog.models import Product, ProductImage
import openpyxl
import os
from django.core.files import File
from pathlib import Path


class Command(BaseCommand):
    help = 'Импорт изображений для товаров из Excel'

    def handle(self, *args, **options):
        excel_path = '/Users/danil/PycharmProjects/Mebel/Расчет_образцов_выставка_Олимпийский_.xlsx'
        images_dir = '/Users/danil/PycharmProjects/Mebel/excel_images'

        self.stdout.write('Загрузка Excel файла...')
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # Словарь для маппинга листов на префиксы изображений
        sheets_mapping = {
            'Двери ': 'Двери',
            'Перегородки': 'Перегородки',
            'Системы хранения': 'Системы хранения',
            'стеллажи': 'стеллажи',
            'витрины': 'витрины',
            'столы': 'столы',
            'диваны': 'диваны',
            'кровати ': 'кровати',
            'полки подвесные ': 'полки подвесные',
            'стеновые панели ': 'стеновые панели',
        }

        total_images_added = 0

        for sheet_name, image_prefix in sheets_mapping.items():
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f'⚠ Лист "{sheet_name}" не найден'))
                continue

            self.stdout.write(f'\n📄 Обработка листа: {sheet_name}')
            ws = wb[sheet_name]

            # Находим строку с заголовками
            header_row = None
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
                if row and any('Наименование товара' in str(cell) for cell in row if cell):
                    header_row = idx
                    break

            if not header_row:
                self.stdout.write(self.style.WARNING(f'  Заголовки не найдены'))
                continue

            # Получаем заголовки
            header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
            headers = [str(cell).strip() if cell else '' for cell in header_cells]

            # Находим индекс колонки с названием товара
            name_col_idx = None
            model_col_idx = None
            for idx, header in enumerate(headers):
                header_lower = header.lower().strip()
                if 'наименование товара' in header_lower:
                    name_col_idx = idx
                elif 'название' in header_lower and ('профиль' in header_lower or 'конструкция' in header_lower):
                    model_col_idx = idx

            # Обрабатываем строки с данными
            row_number = 1  # Счетчик для маппинга изображений
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 2, values_only=True), header_row + 2):
                if not row or not any(row):
                    continue

                # Пропускаем заголовочные строки
                first_cell = str(row[0]).strip() if row[0] else ''
                if not first_cell or first_cell.lower() in ['двери', 'перегородки', 'гардеробные', 'стеллажи']:
                    continue

                # Получаем название товара
                base_name_val = row[name_col_idx] if name_col_idx is not None else None
                base_name = str(base_name_val).strip() if base_name_val and base_name_val != 'None' else ''

                model_name_val = row[model_col_idx] if model_col_idx is not None else None
                model_name = str(model_name_val).strip() if model_name_val and model_name_val != 'None' else ''

                # Формируем название как в import_correct
                if sheet_name in ['Двери ', 'Перегородки', 'столы']:
                    if base_name and model_name:
                        product_name = f'{base_name} {model_name}'
                    else:
                        product_name = base_name
                else:
                    product_name = base_name

                if not product_name:
                    continue

                # Ищем товар в базе
                try:
                    products = Product.objects.filter(name=product_name)
                    if products.count() > 1:
                        # Если несколько товаров с таким именем, берем первый
                        product = products.first()
                    elif products.count() == 1:
                        product = products.first()
                    else:
                        self.stdout.write(self.style.WARNING(f'  ⚠ Товар не найден: {product_name}'))
                        row_number += 1
                        continue

                    # Добавляем изображения
                    images_added = self.add_images(product, image_prefix, row_number, images_dir)
                    if images_added > 0:
                        self.stdout.write(self.style.SUCCESS(f'  ✓ {product.name}: {images_added} изображений'))
                        total_images_added += images_added

                    row_number += 1

                except Exception as e:
                    self.stdout.write(self.style.ERROR(f'  ✗ Ошибка для {product_name}: {str(e)}'))
                    row_number += 1
                    continue

        self.stdout.write(self.style.SUCCESS(f'\n✓ Всего добавлено: {total_images_added} изображений'))

    def add_images(self, product, image_prefix, row_number, images_dir):
        """Добавляет изображения к товару"""
        # Удаляем старые изображения
        product.images.all().delete()

        if not os.path.exists(images_dir):
            return 0

        images_found = []

        # Ищем изображения с паттерном: Префикс_НомерСтроки.расширение
        for file in os.listdir(images_dir):
            if file.startswith('.'):
                continue

            # Проверяем паттерн имени файла
            if file.startswith(f'{image_prefix}_{row_number}.'):
                images_found.append(file)

        # Сортируем файлы
        images_found.sort()

        # Добавляем изображения
        added_count = 0
        for idx, image_file in enumerate(images_found):
            image_path = os.path.join(images_dir, image_file)
            try:
                with open(image_path, 'rb') as f:
                    ProductImage.objects.create(
                        product=product,
                        image=File(f, name=image_file),
                        is_main=(idx == 0),
                        order=idx,
                        alt_text=product.name
                    )
                    added_count += 1
            except Exception as e:
                self.stdout.write(self.style.WARNING(f'    Ошибка загрузки {image_file}: {str(e)}'))

        return added_count
